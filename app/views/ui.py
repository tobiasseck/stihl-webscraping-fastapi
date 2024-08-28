from openpyxl import Workbook
from openpyxl.styles import numbers
import json
import gspread
from google.oauth2.service_account import Credentials

class View:
    @staticmethod
    def save_to_excel(filename, categories, products):
        wb = Workbook()
        wb.remove(wb.active)

        for category in categories:
            ws = wb.create_sheet(title=category.name[:31])
            ws.append(['Name', 'Variant Name', 'Variant Identifier', 'Price', 'Description', 'SKU', 'Link'])

            column_widths = [40, 30, 40, 15, 80, 15, 100]
            for i, width in enumerate(column_widths):
                ws.column_dimensions[chr(65 + i)].width = width

            category_products = [p for p in products if p.category_id == category.id]
            for product in category_products:
                for variant in product.variants:
                    ws.append([
                        product.name,
                        variant.name,
                        variant.identifier,
                        variant.price_history[-1].price if variant.price_history else None,
                        product.description,
                        variant.sku,
                        product.link
                    ])

            for cell in ws['D'][1:]:
                if cell.value is not None:
                    cell.number_format = '€#,##0.00'

        if len(wb.sheetnames) == 0:
            ws = wb.create_sheet(title="No Data")
            ws.append(["No data available"])

        wb.save(filename)

    @staticmethod
    def save_to_json(filename, products):
        data = []
        for product in products:
            for variant in product.variants:
                data.append({
                    "name": product.name,
                    "variant_name": variant.name,
                    "variant_identifier": variant.identifier,
                    "price": variant.price_history[-1].price if variant.price_history else None,
                    "description": product.description,
                    "sku": variant.sku,
                    "link": product.link,
                    "category": product.category.name
                })
        
        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=4)

    @staticmethod
    def save_to_google_sheet(sheet_id, products):
        creds = Credentials.from_service_account_file('stihl-data-efa1a03e0e4f.json', scopes=['https://www.googleapis.com/auth/spreadsheets'])
        client = gspread.authorize(creds)

        sheet = client.open_by_key(sheet_id).sheet1

        headers = ['Name', 'Variant Name', 'Variant Identifier', 'Price', 'Description', 'SKU', 'Link', 'Category']
        data = [headers]

        for product in products:
            for variant in product.variants:
                data.append([
                    product.name,
                    variant.name,
                    variant.identifier,
                    variant.price_history[-1].price if variant.price_history else None,
                    product.description,
                    variant.sku,
                    product.link,
                    product.category.name
                ])

        sheet.clear()
        sheet.update('A1', data)

        sheet.format('A1:H1', {
            "backgroundColor": {
                "red": 0.8,
                "green": 0.8,
                "blue": 0.8
            },
            "textFormat": {"bold": True}
        })

        sheet.columns_auto_resize(0, 8)

        print(f"Data exported to Google Sheet: https://docs.google.com/spreadsheets/d/{sheet_id}")